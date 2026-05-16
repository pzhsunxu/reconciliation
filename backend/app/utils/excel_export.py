from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.models import ReconciliationReport, Hotel, ReconciliationJob
from sqlalchemy.orm import Session


def export_report_to_excel(db: Session, report_id: int) -> bytes:
    """导出对账报表为Excel文件"""
    report = db.query(ReconciliationReport).filter(ReconciliationReport.id == report_id).first()
    if not report:
        return b""

    hotel = db.query(Hotel).filter(Hotel.id == report.hotel_id).first()
    job = db.query(ReconciliationJob).filter(ReconciliationJob.id == report.job_id).first()

    wb = Workbook()
    ws = wb.active
    ws.title = "对账报表"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")

    headers = ["项目", "金额(元)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    data = [
        ("对账周期", f"{report.period_start} 至 {report.period_end}"),
        ("对账任务类型", "自动" if job.job_type == "auto" else "手动"),
        ("酒店名称", hotel.name if hotel else ""),
        ("合作模式", "分润" if hotel and hotel.cooperation_type == "split" else "全租"),
        ("总销售额", f"{report.total_sales:.2f}"),
        ("平台佣金", f"{report.total_commission:.2f}"),
        ("净收入", f"{report.total_net_income:.2f}"),
        ("日常开支", f"{report.total_expense:.2f}"),
        ("可分配利润", f"{report.total_net_income - report.total_expense:.2f}"),
        (f"公司应得(公司占比{(hotel.company_share if hotel else 0) * 100:.0f}%)", f"{report.company_amount:.2f}"),
        (f"房东应得(房东占比{(hotel.owner_share if hotel else 0) * 100:.0f}%)", f"{report.owner_amount:.2f}"),
        ("报表状态", "已复核" if report.status == "reviewed" else "草稿"),
        ("复核人", report.reviewer or ""),
        ("复核时间", str(report.reviewed_at) if report.reviewed_at else ""),
    ]

    for row in data:
        ws.append(list(row) if len(row) > 2 else [*row, ""])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = center_align

    for col in ["A", "B"]:
        ws.column_dimensions[col].width = 20

    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
